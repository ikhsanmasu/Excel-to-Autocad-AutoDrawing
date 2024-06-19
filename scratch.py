from pyautocad import Autocad, APoint
import openpyxl

#librari bounding box
import array as ar
from ctypes import *
from comtypes import automation

def start(path, lebarCellIT1, settingIT1):
    acad = Autocad(create_if_not_exists=True)
    #acad.prompt ini memancing Autocad agar terbuka jika belum terbuka
    acad.prompt("Initialization AutomationBot")

    def convert2list(IT, startRow):
        max = 0
        for i in range(1000):
            if IT.cell(row=i + startRow, column=1).value != None:
                max = i + 1
        convert = lambda x: '' if x == None else x
        return [[convert(IT.cell(row=baris + startRow, column=kolom + 1).value) for kolom in range(26)] for baris in
                range(max)]

    IT = openpyxl.load_workbook(path)
    IT1 = convert2list(IT["IT 1"], 4)
    IT2 = convert2list(IT["IT 2"], 3)

    #membuat IT1
    startX = 0
    startY = 200
    cell = lebarCellIT1
    set = settingIT1
    panjangTabel = sum(cell)

    def tambahBaris(x, y, panjang):
        p1 = APoint(startX + x, startY - y)
        p2 = APoint(startX + x + panjang, startY - y)
        acad.model.AddLine(p1, p2)

    def tambahKolom(x, y, panjang):
        p1 = APoint(startX + x, startY - y)
        p2 = APoint(startX + x, startY - y - panjang)
        acad.model.AddLine(p1, p2)

    # x, y = point pojok kiri atas pada kotak cell
    # lebarCell => lebar bounding box text akan menyesuaikan lebar cell
    def tambahText(x, y, text, besarText, lebarCell, tinggiCell, alignmentCell = "MIDDLE", alignmentText = "CENTER"):
        pCell = APoint(startX + x, startY - y)
        textAcad = acad.model.AddMtext(pCell, lebarCell, text)
        textAcad.height = str(besarText)

        def getBoundingBox():
            #mendapatkan tinggi bounding box text
            # Create 3-d 'Variant' array of 'd'-ouble
            A = automation.VARIANT(ar.array('d', [0, 0, 0]))
            B = automation.VARIANT(ar.array('d', [0, 0, 0]))
            # Get the reference / address
            vA = byref(A)
            vB = byref(B)
            # Call the method from COM object
            textAcad.GetBoundingBox(vA, vB)
            # Return two points as 3-d
            return A.value, B.value

        start, end = getBoundingBox()
        tinggiText = end[1] - start[1]

        #jika tidak middle dan bottom aligment cell adalah TOP
        if alignmentCell == "MIDDLE":
            pTengah = APoint(startX + x, startY - y - tinggiCell / 2 + tinggiText / 2)
            textAcad.move(pCell, pTengah)
        elif alignmentCell == "BOTTOM":
            pTengah = APoint(startX + x, startY - y - tinggiCell + tinggiText)
            textAcad.move(pCell, pTengah)

        if alignmentText == "LEFT" :
            textAcad.AttachmentPoint = "1" #aktualnya TOP LEFT
        elif alignmentText == "CENTER" :
            textAcad.AttachmentPoint = "2" #aktualnya TOP CENTER

        awal, _ = getBoundingBox()
        return startY - y - awal[1]

    def addSimbolR(x, y, lebarCell, tinggiCell, r=1.25):
        x = startX + x + lebarCell/2
        y = startY - y - tinggiCell/2
        p1 = APoint(x, y)
        p2 = APoint(x - r, y)
        p3 = APoint(x + r, y)
        acad.model.AddCircle(p1, r)
        acad.model.AddLine(p2, p3)

    def addSimbolY(x, y, lebarCell, tinggiCell, r=1.25):
        x = startX + x + lebarCell/2
        y = startY - y - tinggiCell/2
        p1 = APoint(x, y)
        p2 = APoint(x - r * 0.7071, y - r * 0.7071)
        p3 = APoint(x + r * 0.7071, y + r * 0.7071)
        acad.model.AddCircle(p1 , r)
        acad.model.AddLine(p2, p3)

    def addSimbolG(x, y, lebarCell, tinggiCell, r=1.25, special=0):
        x = startX + x + lebarCell/2
        y = startY - y - tinggiCell/2
        p1 = APoint(x, y)
        p2 = APoint(x, y + r)
        p3 = APoint(x, y - r)
        acad.model.AddCircle(p1, r)
        acad.model.AddLine(p2, p3)

        if special:
            p1 = APoint(x + r + 0.1, y - r + set[4])
            textAcad = acad.model.AddMtext(p1, lebarCell, "*2")
            textAcad.height = str(set[4])

    def addSimbolE(x, y, lebarCell, tinggiCell, r=1.25):
        x = startX + x + lebarCell/2
        y = startY - y - tinggiCell/2
        p1 = APoint(x, y + r)
        p2 = APoint(x + r, y - r)
        acad.model.AddLine(p1, p2)
        p1 = APoint(x + r, y - r)
        p2 = APoint(x - r, y - r)
        acad.model.AddLine(p1, p2)
        p1 = APoint(x - r, y - r)
        p2 = APoint(x , y + r)
        acad.model.AddLine(p1, p2)

        p1 = APoint(x + r + 0.1, y - r + set[4])
        textAcad = acad.model.AddMtext(p1, lebarCell, "*3")
        textAcad.height = str(set[4])


    def addSimbolL(x, y, lebarCell, tinggiCell, r=1.25):
        x = startX + x + lebarCell/2
        y = startY - y - tinggiCell/2
        p1 = APoint(x - r, y + r)
        p2 = APoint(x + r, y + r)
        acad.model.AddLine(p1, p2)
        p1 = APoint(x + r, y + r)
        p2 = APoint(x + r, y - r)
        acad.model.AddLine(p1, p2)
        p1 = APoint(x + r, y - r)
        p2 = APoint(x - r, y - r)
        acad.model.AddLine(p1, p2)
        p1 = APoint(x - r, y - r)
        p2 = APoint(x - r, y + r)
        acad.model.AddLine(p1, p2)

        p1 = APoint(x - 0.6 * r, y - 0.6 * r)
        acad.model.AddCircle(p1, r/5)
        p1 = APoint(x + 0.6 * r, y + 0.6 * r)
        acad.model.AddCircle(p1, r/5)

    def addSimbolS(x, y, lebarCell, tinggiCell, kecepatan = 0, r=1.5):
        x = startX + x + lebarCell / 2
        y = startY - y - tinggiCell / 2
        p1 = APoint(x, y + r)
        p2 = APoint(x + r, y)
        acad.model.AddLine(p1, p2)
        p1 = APoint(x + r, y)
        p2 = APoint(x, y - r)
        acad.model.AddLine(p1, p2)
        p1 = APoint(x, y - r)
        p2 = APoint(x - r, y)
        acad.model.AddLine(p1, p2)
        p1 = APoint(x - r, y)
        p2 = APoint(x, y + r)
        acad.model.AddLine(p1, p2)
        p1 = APoint(x - r / 2, y + r / 2)
        text = acad.model.AddMText(p1, r, kecepatan)
        text.AttachmentPoint = "2"  # Top Left
        text.height = str(r)  # text height


    #membuat tabel
    tambahBaris(0, 0, panjangTabel)
    tambahBaris(sum(cell[0:2]), set[2] * 0.5, sum(cell[2:25]))
    tambahBaris(sum(cell[0:3]), set[2] * 0.75, sum(cell[3:8]))
    tambahBaris(sum(cell[0:10]), set[2] * 0.75, sum(cell[10:12]))
    tambahBaris(sum(cell[0:13]), set[2] * 0.75, sum(cell[13:15]))
    for index in range(set[0] + 1):
        tambahBaris(0, set[2] + index * set[1], panjangTabel)
    for start in [0, cell[0], sum(cell[0:1]), sum(cell[0:2]), sum(cell[0:12]), sum(cell[0:15]),
                  sum(cell[0:18]), sum(cell[0:25]), panjangTabel]:
        tambahKolom(start, 0, set[2] + set[0] * set[1])
    for start in [sum(cell[0:3]), sum(cell[0:8]), sum(cell[0:9]), sum(cell[0:10]), sum(cell[0:13]),
                  sum(cell[0:16]), sum(cell[0:17]), sum(cell[0:19]), sum(cell[0:20]), sum(cell[0:21]),
                  sum(cell[0:22]), sum(cell[0:23]), sum(cell[0:24])]:
        tambahKolom(start, set[2] * 0.5, set[2] + set[0] * set[1] - set[2] * 0.5)
    for start in [sum(cell[0:4]), sum(cell[0:5]), sum(cell[0:6]), sum(cell[0:7]), sum(cell[0:8]),
                  sum(cell[0:11]), sum(cell[0:14])]:
        tambahKolom(start, set[2] * 0.75, set[2] + set[0] * set[1] - set[2] * 0.75)

    #mengisi Text
    tambahText(0, 0, "ROUTE NO.", set[3], cell[0], set[2])
    bawahRouteName = tambahText(cell[0], 0, "ROUTE\nNAME", set[3], cell[1], set[2])
    tambahText(cell[0], bawahRouteName + 1, "*1", set[4], cell[1], set[2], alignmentCell="TOP")
    tambahText(sum(cell[0:25]), 0, "REMARK", set[3], cell[25], set[2])

    tambahText(sum(cell[0:2]), 0, "SIGNAL AT START OF ROUTE", set[3], sum(cell[2:12]), set[2] * 0.5)
    tambahText(sum(cell[0:12]), 0, "DISTANT SIGNAL", set[3], sum(cell[12:15]), set[2] * 0.5)
    tambahText(sum(cell[0:15]), 0, "DESTINATION OF ROUTE", set[3], sum(cell[15:18]), set[2] * 0.5)
    tambahText(sum(cell[0:18]), 0, "ROUTE CONTROL", set[3], sum(cell[18:25]), set[2] * 0.5)

    #
    tambahText(sum(cell[0:2]), set[2] * 0.5, "NO.", set[3], sum(cell[2:3]), set[2] * 0.5)

    bawahSPIND = tambahText(sum(cell[0:8]), set[2] * 0.5, "SP.\nIND.", set[3], sum(cell[8:9]), set[2] * 0.5)
    tambahText(sum(cell[0:8]), set[2] * 0.5 + bawahSPIND + 0.5, "*4", set[4], sum(cell[8:9]), set[2] * 0.5,
               alignmentCell="TOP")

    tambahText(sum(cell[0:9]), set[2] * 0.5, "CF.\nIND.", set[3], sum(cell[9:10]), set[2] * 0.5)
    tambahText(sum(cell[0:12]), set[2] * 0.5, "NO.", set[3], sum(cell[12:13]), set[2] * 0.5)
    bawahNo = tambahText(sum(cell[0:15]), set[2] * 0.5, "NO.", set[3], sum(cell[15:16]), set[2] * 0.5)
    tambahText(sum(cell[0:15]), set[2] * 0.5 + bawahNo + 1, "*5", set[4], sum(cell[15:16]), set[2] * 0.5,
               alignmentCell="TOP")
    tambahText(sum(cell[0:16]), set[2] * 0.5, "STATION\nNAME", set[3], sum(cell[16:17]), set[2] * 0.5)
    tambahText(sum(cell[0:17]), set[2] * 0.5, "ASP.\nPROV.", set[3], sum(cell[17:18]), set[2] * 0.5)
    tambahText(sum(cell[0:18]), set[2] * 0.5, "POINT LOCKED", set[3], sum(cell[18:19]),
               set[2] * 0.5)
    bawahKeyDetect = tambahText(sum(cell[0:19]), set[2] * 0.5, "KEY\nDETECT", set[3], sum(cell[19:20]),
                                set[2] * 0.5)
    tambahText(sum(cell[0:19]), set[2] * 0.5 + bawahKeyDetect + 0.5, "*7", set[4], sum(cell[19:20]),
               set[2] * 0.5, alignmentCell="TOP")
    bawahTextCC = tambahText(sum(cell[0:20]), set[2] * 0.5, "TRACK CIRCUIT\nCLEAR", set[3], sum(cell[20:21]),
                             set[2] * 0.5)
    tambahText(sum(cell[0:20]), set[2] * 0.5 + bawahTextCC + 0.5, "*8", set[4], sum(cell[20:21]),
               set[2] * 0.5, alignmentCell="TOP")
    tambahText(sum(cell[0:21]), set[2] * 0.5, "SHUNT SIG.\nCLEAR", set[3], sum(cell[21:22]), set[2] * 0.5)
    tambahText(sum(cell[0:22]), set[2] * 0.5, "OPPOSING SIG.\nLOCKED", set[3], sum(cell[22:23]),
               set[2] * 0.5)
    tambahText(sum(cell[0:23]), set[2] * 0.5, "APPROACH\nTRACK", set[3], sum(cell[23:24]), set[2] * 0.5)
    tambahText(sum(cell[0:24]), set[2] * 0.5, "REQ. APPROACH\nTRACK OCCUPIED", set[3], sum(cell[24:25]),
               set[2] * 0.5)
    tambahText(sum(cell[0:3]), set[2] * 0.5, "ASPECT", set[3], sum(cell[3:8]), set[2] * 0.25)
    tambahText(sum(cell[0:10]), set[2] * 0.5, "DIRECTION IND", set[3], sum(cell[10:12]), set[2] * 0.25)
    tambahText(sum(cell[0:13]), set[2] * 0.5, "ASPECT", set[3], sum(cell[13:15]), set[2] * 0.25)

    tambahText(sum(cell[0:10]), set[2] * 0.75, "LEFT", set[3], sum(cell[10:11]), set[2] * 0.25)
    tambahText(sum(cell[0:11]), set[2] * 0.75, "RIGHT", set[3], sum(cell[11:12]), set[2] * 0.25)

    addSimbolR(sum(cell[0:3]), set[2] * 0.75, cell[3], set[2] * 0.25, set[5])
    addSimbolY(sum(cell[0:4]), set[2] * 0.75, cell[4], set[2] * 0.25, set[5])
    addSimbolG(sum(cell[0:5]), set[2] * 0.75, cell[5], set[2] * 0.25, set[5], 1)
    addSimbolE(sum(cell[0:6]), set[2] * 0.75, cell[6], set[2] * 0.25, set[5])
    addSimbolL(sum(cell[0:7]), set[2] * 0.75, cell[7], set[2] * 0.25, set[5])
    addSimbolY(sum(cell[0:13]), set[2] * 0.75, cell[13], set[2] * 0.25, set[5])
    addSimbolG(sum(cell[0:14]), set[2] * 0.75, cell[14], set[2] * 0.25, set[5])
    #

