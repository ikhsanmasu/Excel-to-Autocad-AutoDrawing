from pyautocad import Autocad, APoint
import openpyxl
import math
import numpy as np
import array as ar
from ctypes import *
from comtypes import automation

def initacad():
    acadinit = Autocad(create_if_not_exists=True)
    acadinit.prompt("HALAMAN INI DIBUKA OLEH ISANBOT\n")

acad = Autocad(create_if_not_exists=True)
acad.prompt("HALAMAN INI DIBUKA OLEH ISANBOT\n")

class interlockingTable():
    def __init__(self, path):
        IT = openpyxl.load_workbook(path)
        self.IT1 = IT["IT 1"]
        self.IT2 = IT["IT 2"]
        # self.verifikasi()

    def getIT1(self):
        max = 0
        for i in range(1000):
            if self.IT1.cell(row=i+4, column=1).value != None:
                max = i + 1
        convert = lambda x: '' if x == None else x
        return [[convert(self.IT1.cell(row=baris+4, column=kolom+1).value) for kolom in range(26)] for baris in range(max)]

    def getIT2(self):
        max = 0
        for i in range(1000):
            if self.IT2.cell(row=i+3, column=1).value != None:
                max = i + 1
        convert = lambda x: '' if x == None else x
        return [[convert(self.IT2.cell(row=baris+3, column=kolom+1).value) for kolom in range(15)] for baris in range(max)]

    def verifikasi(self):
        for index in range(self.IT1.max_row - 4):
            startrow = index + 4
            # Cek kolom NO
            NO = self.IT1.cell(row=startrow, column=1).value
            if NO == None:
                print("IT1 A", index + 4, ":: kosong")
            elif type(NO) != int:
                print("IT1 A", index + 4, ":: harus angka")
            elif NO != index + 1:
                print("IT1 A", index + 4, ":: urutan tidak sesuai")

            # Cek kolom Route Name
            if self.IT1.cell(row=startrow, column=2).value == None:
                print("IT1 B", index + 4, ":: kosong")

            # Cek kolom SIGNAL AT START OF ROUTE, NO
            if self.IT1.cell(row=startrow, column=3).value == None:
                print("IT1 C", index + 4, ":: kosong")

            # Cek kolom DESTINATION OF ROUTE, NO
            if self.IT1.cell(row=startrow, column=16).value == None:
                print("IT1 P", index + 4, ":: kosong")

        for index in range(self.IT2.max_row - 4):
            startrow = index + 3
            # Cek kolom NO
            NO = self.IT1.cell(row=startrow, column=1).value
            if NO == None:
                print("IT1 A", index + 4, ":: kosong")
            elif type(NO) != int:
                print("IT1 A", index + 4, ":: harus angka")
            elif NO != index + 1:
                print("IT1 A", index + 4, ":: urutan tidak sesuai")

            # Cek kolom Route Name
            if self.IT1.cell(row=startrow, column=2).value == None:
                print("IT1 B", index + 4, ":: kosong")

            # Cek kolom SIGNAL AT START OF ROUTE, NO
            if self.IT1.cell(row=startrow, column=3).value == None:
                print("IT1 C", index + 4, ":: kosong")

            # Cek kolom DESTINATION OF ROUTE, NO
            if self.IT1.cell(row=startrow, column=16).value == None:
                print("IT1 P", index + 4, ":: kosong")

class ITtoCad(interlockingTable):
    def __init__(self,path):
        super().__init__(path)
        self.x = 0
        self.y = 200

    def addBaris(self, kordinatX, kordinatY,panjang):
        p1 = APoint(self.x + kordinatX, self.y - kordinatY)
        p2 = APoint(self.x + kordinatX + panjang, self.y - kordinatY)
        acad.model.AddLine(p1, p2)

    def addKolom(self, kordinatX, kordinatY,panjang):
        p1 = APoint(self.x + kordinatX, self.y - kordinatY)
        p2 = APoint(self.x + kordinatX, self.y - kordinatY - panjang)
        acad.model.AddLine(p1, p2)

    def GetBoundingBox(self, entity):
        # Create 3-d 'Variant' array of 'd'-ouble
        A = automation.VARIANT(ar.array('d', [0, 0, 0]))
        B = automation.VARIANT(ar.array('d', [0, 0, 0]))
        # Get the reference / address
        vA = byref(A)
        vB = byref(B)
        # Call the method from COM object
        entity.GetBoundingBox(vA, vB)
        # Return two points as 3-d
        return (A.value, B.value)

    def addText(self, x, y, text, tinggicell, lebarcell, textsize=1.5, justify = "MIDDLE CENTER", specialcase = ""):
        if justify == "MIDDLE CENTER":
            p1 = APoint(self.x + x, self.y - y)
            textRouteMC = acad.model.AddMtext(p1, lebarcell, text)  # defined width = self.wRouteNo
            textRouteMC.AttachmentPoint = "2"  # Top center
            textRouteMC.height = str(textsize)  # text height
            start, end = self.GetBoundingBox(textRouteMC)
            p2 = APoint(self.x + x, self.y - y - tinggicell/2 + (self.y - y - start[1]) / 2)
            textRouteMC.move(p1, p2)
            if specialcase != "":
                start, end = self.GetBoundingBox(textRouteMC)
                p3 = APoint(self.x + x, start[1] - 0.5)
                textRouteSP = acad.model.AddMtext(p3, lebarcell, specialcase)
                textRouteSP.AttachmentPoint = "2"  # Top center
                textRouteSP.height = str(1)  # text height

        elif justify == "BOTOM RIGHT":
            tinggi = tinggicell - textsize
            p1 = APoint(self.x + x - 0.5, self.y - y + 0.5 - tinggi)
            textRouteBR = acad.model.AddMtext(p1, lebarcell, text)  # defined width = self.wRouteNo
            textRouteBR.AttachmentPoint = "3"  # BOTOM RIGHT
            textRouteBR.height = str(textsize)  # text height

        elif justify == "TOP LEFT":
            p1 = APoint(self.x + x + 1 , self.y - y)
            textRouteTL = acad.model.AddMtext(p1, lebarcell-2, text)  # defined width = self.wRouteNo
            textRouteTL.AttachmentPoint = "1"  # Top Left
            textRouteTL.height = str(textsize)  # text height
            start, end = self.GetBoundingBox(textRouteTL)
            p2 = APoint(self.x + x + 1, self.y - y - tinggicell / 2 + (self.y - y - start[1]) / 2)
            textRouteTL.move(p1, p2)

    def addSimbolR(self, x, y, lebarcell, tinggicell, r = 1):
        x = self.x + x + lebarcell/2
        y = self.y - y - tinggicell/2
        p1 = APoint(x, y)
        p2 = APoint(x- 1, y)
        p3 = APoint(x+ 1,y)
        acad.model.AddCircle(p1 , r)
        acad.model.AddLine(p2, p3)

    def addSimbolY(self, x, y, lebarcell, tinggicell, r = 1):
        x = self.x + x + lebarcell/2
        y = self.y - y - tinggicell/2
        p1 = APoint(x, y)
        p2 = APoint(x - 0.7071, y - 0.7071)
        p3 = APoint(x + 0.7071, y + 0.7071)
        acad.model.AddCircle(p1 , r)
        acad.model.AddLine(p2, p3)

    def addSimbolG(self, x, y, lebarcell, tinggicell,r = 1):
        x = self.x + x +lebarcell/2
        y = self.y - y - tinggicell/2
        p1 = APoint(x, y)
        p2 = APoint(x, y + 1)
        p3 = APoint(x, y - 1)
        acad.model.AddCircle(p1 , r)
        acad.model.AddLine(p2, p3)

    def addSimbolE(self, x, y, lebarcell, tinggicell,r = 1):
        x = self.x + x +lebarcell/2
        y = self.y - y - tinggicell/2
        p1 = APoint(x, y + r)
        p2 = APoint(x + r, y - r)
        acad.model.AddLine(p1, p2)
        p1 = APoint(x + r, y - r)
        p2 = APoint(x - r, y - r)
        acad.model.AddLine(p1, p2)
        p1 = APoint(x - r, y - r)
        p2 = APoint(x , y + r)
        acad.model.AddLine(p1, p2)

    def addSimbolL(self, x, y, lebarcell, tinggicell, r = 1):
        x = self.x + x + lebarcell/2
        y = self.y - y - tinggicell/2
        p1 = APoint(x  - r, y + r)
        p2 = APoint(x  + r, y + r)
        acad.model.AddLine(p1, p2)
        p1 = APoint(x  + r, y + r)
        p2 = APoint(x  + r, y - r)
        acad.model.AddLine(p1, p2)
        p1 = APoint(x  + r, y - r)
        p2 = APoint(x  - r, y - r)
        acad.model.AddLine(p1, p2)
        p1 = APoint(x  - r, y - r)
        p2 = APoint(x  - r, y + r)
        acad.model.AddLine(p1, p2)

        p1 = APoint(x - 0.6 * r, y - 0.6 * r)
        acad.model.AddCircle(p1, r/5)
        p1 = APoint(x + 0.6 * r, y + 0.6 * r)
        acad.model.AddCircle(p1, r/5)

    def addSimbolS(self, x, y, lebarcell, tinggicell, kecepatan = 0, r = 1.5):
        x = self.x + x + lebarcell / 2
        y = self.y - y - tinggicell / 2
        p1 = APoint(x, y + r)
        p2 = APoint(x + r, y)
        acad.model.AddLine(p1, p2)
        p1 = APoint(x + r, y)
        p2 = APoint(x, y - r)
        acad.model.AddLine(p1, p2)
        p1 = APoint(x, y - r)
        p2 = APoint(x - r, y)
        acad.model.AddLine(p1, p2)
        p1 = APoint(x - r, y)
        p2 = APoint(x, y + r)
        acad.model.AddLine(p1, p2)
        p1 = APoint(x - r / 2, y + r / 2)
        text = acad.model.AddMText(p1, r, kecepatan)
        text.AttachmentPoint = "2"  # Top Left
        text.height = str(r)  # text height


class acadIT1(ITtoCad):
    def __init__(self, path):
        super().__init__(path)
        self.totalbaris = 20
        self.tinggibaris = 7
        self.hHeader = 15

        self.wRouteNo = 8
        self.wRouteName = 15
        self.wSignalAtStart0fRoute_No = 8
        self.wSignalAtStart0fRoute_Aspect_R = 6
        self.wSignalAtStart0fRoute_Aspect_Y = 6
        self.wSignalAtStart0fRoute_Aspect_G = 6
        self.wSignalAtStart0fRoute_Aspect_E = 6
        self.wSignalAtStart0fRoute_Aspect_L = 6
        self.wSignalAtStart0fRoute_S = 6
        self.wSignalAtStart0fRoute_CF = 6
        self.wSignalAtStart0fRoute_Dir_L = 8
        self.wSignalAtStart0fRoute_Dir_R = 8
        self.wDistantSignal_No = 7
        self.wDistantSignal_Aspect_Y = 6
        self.wDistantSignal_Aspect_G = 6
        self.wDestinationOfRoute_No = 10
        self.wDestinationOfRoute_StationName = 10
        self.wDestinationOfRoute_Asp = 10
        self.wRouteControl_Point = 27
        self.wRouteControl_Key = 9
        self.wRouteControl_Track = 42
        self.wRouteControl_Shunt = 13
        self.wRouteControl_OpposingSignall = 17
        self.wRouteControl_ApproachTrack = 13
        self.wRouteControl_ReqApproach = 20
        self.wRemark = 30

        self.wlist =[self.wRouteNo, self.wRouteName, self.wSignalAtStart0fRoute_No,
                     self.wSignalAtStart0fRoute_Aspect_R, self.wSignalAtStart0fRoute_Aspect_Y, self.wSignalAtStart0fRoute_Aspect_G,
                     self.wSignalAtStart0fRoute_Aspect_E, self.wSignalAtStart0fRoute_Aspect_L,
                     self.wSignalAtStart0fRoute_S, self.wSignalAtStart0fRoute_CF,
                     self.wSignalAtStart0fRoute_Dir_L, self.wSignalAtStart0fRoute_Dir_R,
                     self.wDistantSignal_No, self.wDistantSignal_Aspect_Y, self.wDistantSignal_Aspect_G,
                     self.wDestinationOfRoute_No, self.wDestinationOfRoute_StationName, self.wDestinationOfRoute_Asp,
                     self.wRouteControl_Point, self.wRouteControl_Key, self.wRouteControl_Track,
                     self.wRouteControl_Shunt, self.wRouteControl_OpposingSignall, self.wRouteControl_ApproachTrack, self.wRouteControl_ReqApproach,
                     self.wRemark]

        self.wSignalAtStart0fRoute_Aspect = sum(self.wlist[3:8])
        self.wSignalAtStart0fRoute_DirectionIND = sum(self.wlist[10:12])
        self.wSignalAtStart0fRoute = sum(self.wlist[2:12])
        self.wDistantSignal_Aspect = self.wDistantSignal_Aspect_Y + self.wDistantSignal_Aspect_G
        self.wDistantSignal = self.wDistantSignal_No + self.wDistantSignal_Aspect
        self.wDestinationOfRoute = sum(self.wlist[15:18])
        self.wRouteControl = sum(self.wlist[18:25])
        self.wtotal = sum(self.wlist)
        self.xcount = (self.wtotal + 150) * math.ceil(len(self.getIT1()) / self.totalbaris)

    def createIT1(self):

        for num in range(math.ceil(len(self.getIT1()) / self.totalbaris)):
            self.addText(sum(self.wlist) / 2 - 25, -17, "INTERLOCKING TABLE (%s)" %(num+1), 0, 50, 2)

            self.addTable(self.totalbaris)
            self.addHeaderText()
            start = num * self.totalbaris
            if range(math.ceil(len(self.getIT1()) / self.totalbaris)) != num or len(self.getIT1()) % self.totalbaris == 0:
                stop = start + self.totalbaris
            else :
                stop = start + len(self.getIT1()) % self.totalbaris
            self.addIsi(self.getIT1(),start,stop)

            self.addText(0, self.hHeader + self.tinggibaris * self.totalbaris, "*1", 20, 20, 1, "TOP LEFT")
            self.addText(4, self.hHeader + self.tinggibaris * self.totalbaris + 7 + 7,
                         "{\\fArial|b1|i0|c0|p34;\LNOTES}\n\n(T) = TRAIN ROUTE\n(E) = EMERGENCY ROUTE\n(S) = SHUNT ROUTE\n(CF) = CONTRA FLOW ROUTE",
                         0, 30, 1, "TOP LEFT")
            self.addText(0, self.hHeader + self.tinggibaris * self.totalbaris + 10, "*2", 20, 20, 1, "TOP LEFT")
            self.addText(4, self.hHeader + self.tinggibaris * self.totalbaris + 17 + 7,
                         "HOME SIGNAL IS CLEARED TO\nGREEN WHEN THE STARTER\nSIGNAL IS GREEN",
                         0, 30, 1, "TOP LEFT")

            self.addText(sum(self.wlist[0:3]), self.hHeader + self.tinggibaris * self.totalbaris, "*3", 20, 20, 1, "TOP LEFT")
            self.addText(sum(self.wlist[0:3]) + 4, self.hHeader + self.tinggibaris * self.totalbaris + 7 + 7,
                         "EMERGENCY SIGNAL ASPECT IS\nCANCELLED AUTOMATICALLY AFTER\n90 SECOND",
                         0, 30, 1, "TOP LEFT")
            self.addText(sum(self.wlist[0:3]), self.hHeader + self.tinggibaris * self.totalbaris + 10, "*4", 20, 20, 1, "TOP LEFT")
            self.addText(sum(self.wlist[0:3]) + 4, self.hHeader + self.tinggibaris * self.totalbaris + 18 + 7,
                         "SPEED INDICATOR TO BE LAMP PROVED WITH YELLOW\nAND GREEN SIGNAL       \n          3  =   VARIABLE SPEED INDICATOR\n              =   FIX SPEED INDICATOR",
                         0, 50, 1, "TOP LEFT")
            self.addSimbolS(sum(self.wlist[0:3])+ 6.75, self.hHeader + self.tinggibaris * self.totalbaris + 23.8, 5, 5, kecepatan=3, r=1)

            self.addText(sum(self.wlist[0:11]), self.hHeader + self.tinggibaris * self.totalbaris, "*5", 20, 20, 1, "TOP LEFT")
            self.addText(sum(self.wlist[0:11]) + 4, self.hHeader + self.tinggibaris * self.totalbaris + 6.75 + 7,
                         "SIGNAL NUMBER MARKED WITH 'X',\nSHOW OPPOSING SIGNAL WHERE NO\nDESTINATION SIGNAL IS AVAILABLE",
                         0, 30, 1, "TOP LEFT")
            self.addText(sum(self.wlist[0:11]), self.hHeader + self.tinggibaris * self.totalbaris + 10, "*6", 20, 20, 1, "TOP LEFT")
            self.addText(sum(self.wlist[0:11]) + 4, self.hHeader + self.tinggibaris * self.totalbaris + 14.5 + 7,
                         "FIX MARKER",
                         0, 50, 1, "TOP LEFT")

            self.addText(sum(self.wlist[0:16]), self.hHeader + self.tinggibaris * self.totalbaris, "*7", 20, 20, 1, "TOP LEFT")
            self.addText(sum(self.wlist[0:16]) + 4, self.hHeader + self.tinggibaris * self.totalbaris + 8 + 7,
                         "RX-N = \"KEY LOCK\" AT UP POSITION AND\nLOCKED\nDX-N = \"DERAILER\" AT UP POSITION AND\nLOCKED",
                         0, 30, 1, "TOP LEFT")
            self.addText(sum(self.wlist[0:16]), self.hHeader + self.tinggibaris * self.totalbaris + 10, "*8", 20, 20, 1, "TOP LEFT")
            self.addText(sum(self.wlist[0:16]) + 4, self.hHeader + self.tinggibaris * self.totalbaris + 15.5 + 7,
                         "{\LUNDERLINE} = TRACK CIRCUIT WITH\nUNDERLINE IS OPPOSITE SIGNAL TRACK",
                         0, 50, 1, "TOP LEFT")

            self.x += sum(self.wlist) + 150

    def addTable(self,baris):
        self.addBaris(0, 0, self.wtotal)
        self.addBaris(sum(self.wlist[0:2]), self.hHeader/2, self.wtotal - self.wRouteNo - self.wRouteName - self.wRemark)
        self.addBaris(sum(self.wlist[0:3]), self.hHeader * 3 / 4, self.wSignalAtStart0fRoute_Aspect)
        self.addBaris(sum(self.wlist[0:10]), self.hHeader * 3 / 4, self.wSignalAtStart0fRoute_DirectionIND)
        self.addBaris(sum(self.wlist[0:13]), self.hHeader * 3 / 4, self.wDistantSignal_Aspect)
        self.addBaris(0, self.hHeader, self.wtotal)

        tinggi = self.tinggibaris
        hTotal = self.hHeader
        jumlahbaris = baris
        for i in range(jumlahbaris):
            hTotal += tinggi
            self.addBaris(0, hTotal, self.wtotal)

        start = 0
        for i in [0, self.wRouteNo, self.wRouteName, self.wSignalAtStart0fRoute, self.wDistantSignal, self.wDestinationOfRoute, self.wRouteControl, self.wRemark]:
            start += i
            self.addKolom(start, 0, hTotal)

        start = 0
        for i in [self.wRouteNo + self.wRouteName + self.wSignalAtStart0fRoute_No, self.wSignalAtStart0fRoute_Aspect, self.wSignalAtStart0fRoute_S, self.wSignalAtStart0fRoute_CF, self.wSignalAtStart0fRoute_DirectionIND + self.wDistantSignal_No, self.wDistantSignal_Aspect + self.wDestinationOfRoute_No, self.wDestinationOfRoute_StationName,
                  self.wDestinationOfRoute_Asp + self.wRouteControl_Point, self.wRouteControl_Key, self.wRouteControl_Track, self.wRouteControl_Shunt, self.wRouteControl_OpposingSignall, self.wRouteControl_ApproachTrack]:
            start += i
            self.addKolom(start, self.hHeader / 2, hTotal - self.hHeader / 2)

        start = 0
        for i in [self.wRouteNo + self.wRouteName + self.wSignalAtStart0fRoute_No + self.wSignalAtStart0fRoute_Aspect_R, self.wSignalAtStart0fRoute_Aspect_Y, self.wSignalAtStart0fRoute_Aspect_G, self.wSignalAtStart0fRoute_Aspect_E, self.wSignalAtStart0fRoute_Aspect_L, self.wSignalAtStart0fRoute_S + self.wSignalAtStart0fRoute_CF + self.wSignalAtStart0fRoute_Dir_L, self.wSignalAtStart0fRoute_Dir_R + self.wDistantSignal_No + self.wDistantSignal_Aspect_G]:
            start += i
            self.addKolom(start, self.hHeader * 3 / 4, hTotal - self.hHeader * 3 / 4)


    def addHeaderText(self):
        self.addText(0, 0, "ROUTE\nNO.", self.hHeader, self.wRouteNo)
        self.addText(self.wRouteNo, 0, "ROUTE\nNAME", self.hHeader, self.wRouteName, specialcase= "*1")
        # self.addText(self.wRouteNo, 0, "*1", self.hHeader + 6, self.wRouteName, 1)
        self.addText(sum(self.wlist[0:2]), 0, "SIGNAL AT START OF ROUTE", self.hHeader / 2, self.wSignalAtStart0fRoute)
        self.addText(sum(self.wlist[0:12]), 0, "DISTANT SIGNAL", self.hHeader / 2, self.wDistantSignal)
        self.addText(sum(self.wlist[0:15]), 0, "DESTINATION OF ROUTE", self.hHeader / 2, self.wDestinationOfRoute)
        self.addText(sum(self.wlist[0:18]), 0, "ROUTE CONTROL", self.hHeader / 2, self.wRouteControl)
        self.addText(sum(self.wlist[0:25]), 0, "REMARK", self.hHeader, self.wRemark)

        self.addText(sum(self.wlist[0:2]), self.hHeader / 2, "NO.", self.hHeader / 2, self.wSignalAtStart0fRoute_No)
        self.addText(sum(self.wlist[0:8]), self.hHeader / 2, "SP.\nIND.", self.hHeader / 2, self.wSignalAtStart0fRoute_S, specialcase = "*4")
        # self.addText(sum(self.wlist[0:8]), self.hHeader / 2 + 3, "*4", self.hHeader / 2, self.wSignalAtStart0fRoute_S, 1)
        self.addText(sum(self.wlist[0:9]),self.hHeader / 2, "CF.\nIND.", self.hHeader / 2, self.wSignalAtStart0fRoute_CF)

        self.addText(sum(self.wlist[0:12]), self.hHeader / 2, "NO.", self.hHeader / 2, self.wDistantSignal_No)
        self.addText(sum(self.wlist[0:15]), self.hHeader / 2, "NO.", self.hHeader / 2, self.wDestinationOfRoute_No, specialcase = "*5")
        # self.addText(sum(self.wlist[0:15]), self.hHeader / 2 + 2.25, "*5", self.hHeader / 2, self.wDestinationOfRoute_No, 1)
        self.addText(sum(self.wlist[0:16]), self.hHeader / 2, "STATION\nNAME", self.hHeader / 2, self.wDestinationOfRoute_StationName)
        self.addText(sum(self.wlist[0:17]), self.hHeader / 2, "ASP.\nPROV.", self.hHeader / 2, self.wDestinationOfRoute_Asp)
        self.addText(sum(self.wlist[0:18]), self.hHeader / 2, "POINTS LOCKED", self.hHeader / 2, self.wRouteControl_Point)
        self.addText(sum(self.wlist[0:19]), self.hHeader / 2, "KEY\nDETECT", self.hHeader / 2, self.wRouteControl_Key,specialcase = "*7")
        # self.addText(sum(self.wlist[0:19]), self.hHeader / 2 + 2.8, "*7", self.hHeader / 2, self.wRouteControl_Key, 1)
        self.addText(sum(self.wlist[0:20]), self.hHeader / 2, "TRACK CIRCUIT\nCLEAR", self.hHeader / 2, self.wRouteControl_Track, specialcase = "*8")
        # self.addText(sum(self.wlist[0:20]) + 4, self.hHeader * 3 / 4, "\n*8", 1.2, self.wRouteControl_Track, 1)
        self.addText(sum(self.wlist[0:21]), self.hHeader / 2, "SHUNT SIG.\nCLEAR", self.hHeader / 2, self.wRouteControl_Shunt)
        self.addText(sum(self.wlist[0:22]), self.hHeader / 2, "OPPOSING SIG.\nLOCKED", self.hHeader / 2, self.wRouteControl_OpposingSignall)
        self.addText(sum(self.wlist[0:23]), self.hHeader / 2, "APPROACH\nTRACK", self.hHeader / 2, self.wRouteControl_ApproachTrack)
        self.addText(sum(self.wlist[0:24]), self.hHeader / 2, "REQ. APPROACH\nTRACK OCCUPIED", self.hHeader / 2, self.wRouteControl_ReqApproach)

        self.addText(sum(self.wlist[0:3]), self.hHeader / 2, "ASPECT", self.hHeader / 4, self.wSignalAtStart0fRoute_Aspect)
        self.addText(sum(self.wlist[0:10]), self.hHeader / 2, "DIRECTION IND", self.hHeader / 4, self.wSignalAtStart0fRoute_DirectionIND)
        self.addText(sum(self.wlist[0:13]), self.hHeader / 2, "ASPECT", self.hHeader / 4, self.wDistantSignal_Aspect)
        self.addText(sum(self.wlist[0:10]), self.hHeader * 3 / 4, "LEFT", self.hHeader / 4, self.wSignalAtStart0fRoute_Dir_L)
        self.addText(sum(self.wlist[0:11]), self.hHeader * 3 / 4, "RIGHT", self.hHeader / 4, self.wSignalAtStart0fRoute_Dir_L)

        self.addSimbolR(sum(self.wlist[0:3]), self.hHeader * 3 / 4, self.wSignalAtStart0fRoute_Aspect_R, self.hHeader / 4)
        self.addSimbolY(sum(self.wlist[0:4]), self.hHeader * 3 / 4, self.wSignalAtStart0fRoute_Aspect_R, self.hHeader / 4)
        self.addSimbolG(sum(self.wlist[0:5]), self.hHeader * 3 / 4, self.wSignalAtStart0fRoute_Aspect_R, self.hHeader / 4)
        self.addText(sum(self.wlist[0:5]), self.hHeader * 3 / 4,"*2", self.hHeader / 4, self.wSignalAtStart0fRoute_Aspect_R, 1, "BOTOM RIGHT")
        self.addSimbolE(sum(self.wlist[0:6]), self.hHeader * 3 / 4, self.wSignalAtStart0fRoute_Aspect_R, self.hHeader / 4)
        self.addText(sum(self.wlist[0:6]), self.hHeader * 3 / 4,"*3", self.hHeader / 4, self.wSignalAtStart0fRoute_Aspect_R, 1, "BOTOM RIGHT")
        self.addSimbolL(sum(self.wlist[0:7]), self.hHeader * 3 / 4, self.wSignalAtStart0fRoute_Aspect_R, self.hHeader / 4)

        self.addSimbolY(sum(self.wlist[0:13]), self.hHeader * 3 / 4, self.wSignalAtStart0fRoute_Aspect_R, self.hHeader / 4)
        self.addSimbolG(sum(self.wlist[0:14]), self.hHeader * 3 / 4, self.wSignalAtStart0fRoute_Aspect_R, self.hHeader / 4)

    def addIsi(self,IT,start,stop):
        pass
        for countdata, data in enumerate(IT[start:stop]):
            for countcolum, colum in enumerate(data):

                if countcolum == 0 and str(colum).count(".") != 1:
                    self.addText(sum(self.wlist[0:countcolum]), self.hHeader + countdata * self.tinggibaris, str(colum) + ".", self.tinggibaris, self.wlist[countcolum])

                elif countcolum == 8 and str(colum).count("#") != 0:
                    kecepatan = int(str(colum).replace("#",""))
                    self.addSimbolS(sum(self.wlist[0:countcolum]), self.hHeader + countdata * self.tinggibaris,  self.wlist[countcolum], self.tinggibaris, kecepatan)

                elif countcolum == 18:
                    self.addText(sum(self.wlist[0:countcolum]), self.hHeader + countdata * self.tinggibaris, str(colum), self.tinggibaris, self.wlist[countcolum],1.15,"TOP LEFT")

                elif countcolum == 20:

                    if str(colum).count("T") <= 1:
                        text = "{\L%s}" % str(colum)
                    else:
                        text = "{\L%s}" % str(colum)[0:str(colum).find(" ")] + str(colum)[str(colum).find(" ")::]

                    self.addText(sum(self.wlist[0:countcolum]), self.hHeader + countdata * self.tinggibaris, text, self.tinggibaris, self.wlist[countcolum],1.15,"TOP LEFT")

                elif countcolum == 16:
                    self.addText(sum(self.wlist[0:countcolum]), self.hHeader + countdata * self.tinggibaris, str(colum), self.tinggibaris, self.wlist[countcolum], 1)

                else:
                    self.addText(sum(self.wlist[0:countcolum]), self.hHeader + countdata * self.tinggibaris, str(colum), self.tinggibaris, self.wlist[countcolum])


class acadIT2(ITtoCad):
    def __init__(self, path):
        super().__init__(path)
        self.totalbaris = 20
        self.tinggibaris = 7
        self.hHeader = 15

        self.wRouteNo = 8
        self.wRouteName = 15.5
        self.wSignalNumber = 11
        self.wRouteFlankProtectionControls_PointLocked = 18
        self.wRouteFlankProtectionControls_SignalLocked = 32
        self.wRouteFlankProtectionControls_TrackClear = 8.5
        self.wOverlap_StartSignal = self.wOverlap = 17
        self.wOverlapControls_PointLocked = 21
        self.wOverlapControls_KeyDetect = 18
        self.wOverlapControls_TrackClear = 19
        self.wOverlapControls_OpposingSignalLocked = 22
        self.wOverlapFlankProtectionControls_PointLocked = 14
        self.wOverlapFlankProtectionControls_SignalLocked = 26
        self.wOverlapFlankProtectionControls_TrackClear = 9
        self.wRemark = 75

        self.wlist = [self.wRouteNo,
                        self.wRouteName,
                        self.wSignalNumber,
                        self.wRouteFlankProtectionControls_PointLocked,
                        self.wRouteFlankProtectionControls_SignalLocked,
                        self.wRouteFlankProtectionControls_TrackClear,
                        self.wOverlap_StartSignal,
                        self.wOverlapControls_PointLocked,
                        self.wOverlapControls_KeyDetect,
                        self.wOverlapControls_TrackClear,
                        self.wOverlapControls_OpposingSignalLocked,
                        self.wOverlapFlankProtectionControls_PointLocked,
                        self.wOverlapFlankProtectionControls_SignalLocked,
                        self.wOverlapFlankProtectionControls_TrackClear,
                        self.wRemark]

        self.wRouteFlankProtectionControls = sum(self.wlist[3:6])
        self.wOverlapControls = sum(self.wlist[7:11])
        self.wOverlapFlankProtectionControls = sum(self.wlist[11:14])

        self.wtotal = sum(self.wlist)

    def createIT2(self, itcount):
        self.x = itcount
        for num in range(math.ceil(len(self.getIT2()) / self.totalbaris)):
            self.addText(sum(self.wlist) / 2 - 25, -17, "INTERLOCKING TABLE (%s)" %(num+1+math.ceil(len(self.getIT1()) / self.totalbaris)), 0, 50, 2)

            self.addTable(self.totalbaris)
            self.addHeaderText()
            start = num * self.totalbaris
            if range(math.ceil(len(self.getIT2()) / self.totalbaris)) != num or len(self.getIT2()) % self.totalbaris == 0:
                stop = start + self.totalbaris
            else :
                stop = start + len(self.getIT1()) % self.totalbaris
            self.addIsi(self.getIT2(), start, stop)

            self.x += sum(self.wlist) + 150

    def addTable(self,jumlahbaris):
        self.addBaris(0, 0, self.wtotal)
        self.addBaris(sum(self.wlist[0:3]), self.hHeader / 2, sum(self.wlist[3:14]))
        self.addBaris(0, self.hHeader, self.wtotal)

        tinggi = self.tinggibaris
        hTotal = self.hHeader
        for i in range(jumlahbaris):
            hTotal += tinggi
            self.addBaris(0, hTotal, self.wtotal)

        start = 0
        for i in [0, self.wRouteNo, self.wRouteName, self.wSignalNumber, self.wRouteFlankProtectionControls, self.wOverlap, self.wOverlapControls, self.wOverlapFlankProtectionControls, self.wRemark]:
            start += i
            self.addKolom(start, 0, hTotal)

        start = 0
        for i in [sum(self.wlist[0:4]), self.wRouteFlankProtectionControls_SignalLocked,
                  self.wRouteFlankProtectionControls_TrackClear + self.wOverlap + self.wOverlapControls_PointLocked, self.wOverlapControls_KeyDetect,
                  self.wOverlapControls_TrackClear, self.wOverlapControls_OpposingSignalLocked + self.wOverlapFlankProtectionControls_PointLocked, self.wOverlapFlankProtectionControls_SignalLocked]:
            start += i
            self.addKolom(start, self.hHeader / 2, hTotal - self.hHeader / 2)

    def addHeaderText(self):
        self.addText(0, 0, "ROUTE\nNO.", self.hHeader, self.wRouteNo)
        self.addText(self.wRouteNo, 0, "ROUTE\nNAME", self.hHeader, self.wRouteName)
        self.addText(self.wRouteNo + self.wRouteName, 0, "SIGNAL\nNUMBER", self.hHeader, self.wSignalNumber)
        self.addText(sum(self.wlist[0:14]), 0, "REMARK", self.hHeader, self.wRemark)

        self.addText(sum(self.wlist[0:3]), 0, "ROUTE FLANK PROTECTION CONTROLS", self.hHeader / 2, self.wRouteFlankProtectionControls)
        self.addText(sum(self.wlist[0:6]), 0, "OVERLAP", self.hHeader / 2, self.wOverlap)
        self.addText(sum(self.wlist[0:7]), 0, "OVERLAP CONTROLS", self.hHeader / 2, self.wOverlapControls)
        self.addText(sum(self.wlist[0:11]), 0, "OVERLAP FLANK PROTECTION CONTROLS", self.hHeader / 2, self.wOverlapFlankProtectionControls)

        self.addText(sum(self.wlist[0:3]), self.hHeader / 2, "POINTS LOCKED", self.hHeader / 2, self.wRouteFlankProtectionControls_PointLocked)
        self.addText(sum(self.wlist[0:4]), self.hHeader / 2, "SIGNAL LOCKED &\nPROVED AT DANGER", self.hHeader / 2,
                     self.wRouteFlankProtectionControls_SignalLocked)
        self.addText(sum(self.wlist[0:5]), self.hHeader / 2, "TRACK CLEAR", self.hHeader / 2,
                     self.wRouteFlankProtectionControls_TrackClear)
        self.addText(sum(self.wlist[0:6]), self.hHeader / 2, "START SIGNAL", self.hHeader / 2,
                     self.wOverlap_StartSignal)
        self.addText(sum(self.wlist[0:7]), self.hHeader / 2, "POINTS LOCKED", self.hHeader / 2,
                     self.wOverlapControls_PointLocked)
        self.addText(sum(self.wlist[0:8]), self.hHeader / 2, "KET\nDETECT", self.hHeader / 2,
                     self.wOverlapControls_KeyDetect)
        self.addText(sum(self.wlist[0:9]), self.hHeader / 2, "TRACK\nCLEAR", self.hHeader / 2,
                     self.wOverlapControls_TrackClear)
        self.addText(sum(self.wlist[0:10]), self.hHeader / 2, "OPPOSING SIGNAL\nLOCKED", self.hHeader / 2,
                     self.wOverlapControls_OpposingSignalLocked)
        self.addText(sum(self.wlist[0:11]), self.hHeader / 2, "POINTS\nLOCKED", self.hHeader / 2,
                     self.wOverlapFlankProtectionControls_PointLocked)
        self.addText(sum(self.wlist[0:12]), self.hHeader / 2, "SIGNAL LOCKED &\nPROVED AT DANGER", self.hHeader / 2,
                     self.wOverlapFlankProtectionControls_SignalLocked)
        self.addText(sum(self.wlist[0:13]), self.hHeader / 2, "TRACK\nCLEAR", self.hHeader / 2,
                     self.wOverlapFlankProtectionControls_TrackClear)

    def addIsi(self, IT, start, stop):
        pass
        pass
        for countdata, data in enumerate(IT[start:stop]):
            for countcolum, colum in enumerate(data):
                self.addText(sum(self.wlist[0:countcolum]), self.hHeader + countdata * self.tinggibaris, str(colum), self.tinggibaris, self.wlist[countcolum])


def acadITcreate(path):
    IT1 = acadIT1(path)
    IT1.x = 0
    IT2 = acadIT2(path)
    IT1.createIT1()
    IT2.createIT2(IT1.xcount)


